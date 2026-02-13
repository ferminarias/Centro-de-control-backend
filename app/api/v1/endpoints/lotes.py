import io
import logging
import uuid

from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import verify_admin_key
from app.models.account import Account
from app.models.field import CustomField
from app.models.lead import Lead
from app.models.lead_base import LeadBase
from app.models.lote import Lote
from app.models.record import Record
from app.schemas.lead import LeadListResponse
from app.schemas.lote import (
    LoteAssociateRequest,
    LoteAssociateResponse,
    LoteListResponse,
    LoteResponse,
)
from app.services.field_auto_creator import auto_create_fields

logger = logging.getLogger(__name__)
router = APIRouter(dependencies=[Depends(verify_admin_key)])


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------
@router.get(
    "/accounts/{account_id}/lotes/template",
    summary="Download Excel template for batch import",
)
def download_template(
    account_id: uuid.UUID,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    fields = (
        db.query(CustomField.nombre_campo)
        .filter(CustomField.cuenta_id == account_id)
        .order_by(CustomField.created_at)
        .all()
    )
    headers = [f[0] for f in fields]

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(headers if headers else ["campo_ejemplo"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"plantilla_{account.nombre}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Import Excel → create Lote + Leads
# ---------------------------------------------------------------------------
@router.post(
    "/accounts/{account_id}/lotes/import",
    response_model=LoteResponse,
    summary="Import leads from Excel file",
)
def import_lote(
    account_id: uuid.UUID,
    file: UploadFile,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
) -> dict:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    # Read the uploaded Excel file
    try:
        content = file.file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file must have a header row and at least one data row")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if not any(header):
        raise HTTPException(status_code=400, detail="Header row is empty")

    # Auto-create fields if enabled
    if account.auto_crear_campos:
        existing_fields = (
            db.query(CustomField.nombre_campo)
            .filter(CustomField.cuenta_id == account.id)
            .all()
        )
        existing_names: set[str] = {f[0] for f in existing_fields}
        dummy_payload = {col: "" for col in header if col}
        auto_create_fields(db, account.id, dummy_payload, existing_names)

    # Create Lote
    lote = Lote(
        cuenta_id=account.id,
        nombre=nombre,
        total_leads=0,
    )
    db.add(lote)
    db.flush()

    # Create leads from rows
    count = 0
    for row in rows[1:]:
        datos = {}
        for i, val in enumerate(row):
            if i < len(header) and header[i]:
                datos[header[i]] = val if val is not None else ""
        if not any(v for v in datos.values() if v != ""):
            continue  # skip empty rows

        record = Record(
            cuenta_id=account.id,
            datos=datos,
            metadata_={"source": "lote_import", "lote_id": str(lote.id)},
        )
        db.add(record)
        db.flush()

        lead = Lead(
            cuenta_id=account.id,
            record_id=record.id,
            datos=datos,
            lote_id=lote.id,
        )
        db.add(lead)
        count += 1

    lote.total_leads = count
    db.commit()
    db.refresh(lote)

    logger.info("Lote '%s' created with %d leads for account %s", nombre, count, account_id)

    return {
        "id": lote.id,
        "cuenta_id": lote.cuenta_id,
        "nombre": lote.nombre,
        "lead_base_id": lote.lead_base_id,
        "base_nombre": None,
        "total_leads": lote.total_leads,
        "created_at": lote.created_at,
    }


# ---------------------------------------------------------------------------
# List lotes
# ---------------------------------------------------------------------------
@router.get(
    "/accounts/{account_id}/lotes",
    response_model=LoteListResponse,
    summary="List lotes for an account",
)
def list_lotes(
    account_id: uuid.UUID,
    db: Session = Depends(get_db),
) -> dict:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    lotes = (
        db.query(Lote)
        .filter(Lote.cuenta_id == account_id)
        .order_by(Lote.created_at.desc())
        .all()
    )

    # Fetch base names
    base_ids = {lo.lead_base_id for lo in lotes if lo.lead_base_id}
    base_names: dict[uuid.UUID, str] = {}
    if base_ids:
        bases = db.query(LeadBase.id, LeadBase.nombre).filter(LeadBase.id.in_(base_ids)).all()
        base_names = {b.id: b.nombre for b in bases}

    items = []
    for lo in lotes:
        items.append({
            "id": lo.id,
            "cuenta_id": lo.cuenta_id,
            "nombre": lo.nombre,
            "lead_base_id": lo.lead_base_id,
            "base_nombre": base_names.get(lo.lead_base_id) if lo.lead_base_id else None,
            "total_leads": lo.total_leads,
            "created_at": lo.created_at,
        })

    return {"items": items, "total": len(items)}


# ---------------------------------------------------------------------------
# Get lote detail
# ---------------------------------------------------------------------------
@router.get(
    "/lotes/{lote_id}",
    response_model=LoteResponse,
    summary="Get lote details",
)
def get_lote(
    lote_id: uuid.UUID,
    db: Session = Depends(get_db),
) -> dict:
    lote = db.query(Lote).filter(Lote.id == lote_id).first()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote not found")

    base_nombre = None
    if lote.lead_base_id:
        base = db.query(LeadBase.nombre).filter(LeadBase.id == lote.lead_base_id).first()
        if base:
            base_nombre = base.nombre

    return {
        "id": lote.id,
        "cuenta_id": lote.cuenta_id,
        "nombre": lote.nombre,
        "lead_base_id": lote.lead_base_id,
        "base_nombre": base_nombre,
        "total_leads": lote.total_leads,
        "created_at": lote.created_at,
    }


# ---------------------------------------------------------------------------
# Delete lote (and its leads)
# ---------------------------------------------------------------------------
@router.delete(
    "/lotes/{lote_id}",
    summary="Delete lote and its leads",
)
def delete_lote(
    lote_id: uuid.UUID,
    db: Session = Depends(get_db),
) -> dict:
    lote = db.query(Lote).filter(Lote.id == lote_id).first()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote not found")

    # Delete leads belonging to this lote (and their records)
    leads = db.query(Lead).filter(Lead.lote_id == lote_id).all()
    record_ids = [l.record_id for l in leads]

    db.query(Lead).filter(Lead.lote_id == lote_id).delete(synchronize_session=False)
    if record_ids:
        db.query(Record).filter(Record.id.in_(record_ids)).delete(synchronize_session=False)

    db.delete(lote)
    db.commit()

    logger.info("Lote %s deleted with %d leads", lote_id, len(leads))
    return {"detail": f"Lote deleted with {len(leads)} leads"}


# ---------------------------------------------------------------------------
# Associate / disassociate lote to a base
# ---------------------------------------------------------------------------
@router.put(
    "/lotes/{lote_id}/associate",
    response_model=LoteAssociateResponse,
    summary="Associate or disassociate lote to a lead base",
)
def associate_lote(
    lote_id: uuid.UUID,
    body: LoteAssociateRequest,
    db: Session = Depends(get_db),
) -> dict:
    lote = db.query(Lote).filter(Lote.id == lote_id).first()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote not found")

    target_base_id = body.lead_base_id

    if target_base_id is not None:
        # Verify base exists and belongs to same account
        base = db.query(LeadBase).filter(
            LeadBase.id == target_base_id,
            LeadBase.cuenta_id == lote.cuenta_id,
        ).first()
        if not base:
            raise HTTPException(status_code=404, detail="Lead base not found or belongs to different account")
    else:
        # Disassociate → move leads to default base of the account
        default_base = db.query(LeadBase).filter(
            LeadBase.cuenta_id == lote.cuenta_id,
            LeadBase.es_default.is_(True),
        ).first()
        target_base_id = default_base.id if default_base else None

    # Update lote
    lote.lead_base_id = body.lead_base_id  # store the actual request value (None if disassociate)

    # Move all leads of this lote to the target base
    leads_moved = (
        db.query(Lead)
        .filter(Lead.lote_id == lote_id)
        .update({"lead_base_id": target_base_id}, synchronize_session=False)
    )

    db.commit()

    logger.info(
        "Lote %s associated to base %s, %d leads moved",
        lote_id, target_base_id, leads_moved,
    )

    return {
        "lote_id": lote.id,
        "lead_base_id": target_base_id,
        "leads_moved": leads_moved,
    }


# ---------------------------------------------------------------------------
# List leads of a lote (paginated)
# ---------------------------------------------------------------------------
@router.get(
    "/lotes/{lote_id}/leads",
    response_model=LeadListResponse,
    summary="List leads in a lote",
)
def list_lote_leads(
    lote_id: uuid.UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    lote = db.query(Lote).filter(Lote.id == lote_id).first()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote not found")

    query = db.query(Lead).filter(Lead.lote_id == lote_id)
    total = query.count()
    leads = (
        query.order_by(Lead.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Fetch base names
    base_ids = {l.lead_base_id for l in leads if l.lead_base_id}
    base_names: dict[uuid.UUID, str] = {}
    if base_ids:
        bases = db.query(LeadBase.id, LeadBase.nombre).filter(LeadBase.id.in_(base_ids)).all()
        base_names = {b.id: b.nombre for b in bases}

    items = []
    for lead in leads:
        items.append({
            "id": lead.id,
            "cuenta_id": lead.cuenta_id,
            "record_id": lead.record_id,
            "lead_base_id": lead.lead_base_id,
            "base_nombre": base_names.get(lead.lead_base_id) if lead.lead_base_id else None,
            "lote_id": lead.lote_id,
            "lote_nombre": lote.nombre,
            "datos": lead.datos,
            "created_at": lead.created_at,
        })

    return {"items": items, "total": total, "page": page, "page_size": page_size}
