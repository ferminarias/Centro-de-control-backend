import io
import json
import logging
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from app.core.auth import get_current_user
from app.core.database import get_db
from app.core.multi_tenant import verify_tenant_access, TenantContext, get_tenant_context
from app.core.security import verify_admin_key
from app.models.account import Account
from app.models.field import CustomField
from app.models.lead import Lead
from app.models.lead_base import LeadBase
from app.models.lote import Lote
from app.models.user import User
from app.schemas.lead import BulkUpdateResponse, LeadListResponse, LeadResponse
from app.utils.column_manager import sync_lead_columns

logger = logging.getLogger(__name__)
router = APIRouter(dependencies=[Depends(verify_admin_key)])


def _lead_to_dict(lead: Lead, base_nombre: str | None = None, lote_nombre: str | None = None) -> dict:
    return {
        "id": lead.id,
        "id_lead": lead.id_lead,
        "cuenta_id": lead.cuenta_id,
        "record_id": lead.record_id,
        "lead_base_id": lead.lead_base_id,
        "base_nombre": base_nombre,
        "lote_id": lead.lote_id,
        "lote_nombre": lote_nombre,
        "datos": lead.datos,
        "created_at": lead.created_at,
    }


@router.get(
    "/accounts/{account_id}/leads",
    response_model=LeadListResponse,
    summary="List leads for an account with advanced filters",
)
def list_leads(
    account_id: uuid.UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    # Search
    search: str | None = Query(None, description="Full-text search on nombre, email, telefono"),
    # Filters
    lead_base_id: uuid.UUID | None = Query(None),
    lote_id: uuid.UUID | None = Query(None),
    assigned_to: uuid.UUID | None = Query(None),
    tipificacion_id: uuid.UUID | None = Query(None),
    temperatura: str | None = Query(None, description="frio, tibio, caliente"),
    score_min: int | None = Query(None, ge=0, le=100),
    score_max: int | None = Query(None, ge=0, le=100),
    created_after: datetime | None = Query(None),
    created_before: datetime | None = Query(None),
    # Sorting
    order_by: str = Query("created_at", description="created_at, score, id_lead"),
    order_dir: str = Query("desc", description="asc or desc"),
    db: Session = Depends(get_db),
) -> dict:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    query = (
        db.query(Lead)
        .options(joinedload(Lead.lead_base), joinedload(Lead.lote))
        .filter(Lead.cuenta_id == account_id)
    )

    # Full-text search across JSONB fields
    if search:
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Lead.datos["nombre"].astext.ilike(term),
                Lead.datos["email"].astext.ilike(term),
                Lead.datos["telefono"].astext.ilike(term),
                Lead.datos["phone"].astext.ilike(term),
                Lead.datos["apellido"].astext.ilike(term),
            )
        )

    if lead_base_id:
        query = query.filter(Lead.lead_base_id == lead_base_id)
    if lote_id:
        query = query.filter(Lead.lote_id == lote_id)
    if assigned_to:
        query = query.filter(Lead.assigned_to == assigned_to)
    if tipificacion_id:
        query = query.filter(Lead.tipificacion_id == tipificacion_id)
    if temperatura:
        query = query.filter(Lead.temperatura == temperatura)
    if score_min is not None:
        query = query.filter(Lead.score >= score_min)
    if score_max is not None:
        query = query.filter(Lead.score <= score_max)
    if created_after:
        query = query.filter(Lead.created_at >= created_after)
    if created_before:
        query = query.filter(Lead.created_at <= created_before)

    # Sorting
    sort_col = {
        "created_at": Lead.created_at,
        "score": Lead.score,
        "id_lead": Lead.id_lead,
    }.get(order_by, Lead.created_at)
    query = query.order_by(sort_col.desc() if order_dir == "desc" else sort_col.asc())

    total = query.count()
    leads = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [
        _lead_to_dict(
            lead,
            lead.lead_base.nombre if lead.lead_base else None,
            lead.lote.nombre if lead.lote else None,
        )
        for lead in leads
    ]

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get(
    "/leads/{lead_id}",
    response_model=LeadResponse,
    summary="Get lead details",
)
def get_lead(
    lead_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    # Verify tenant access - lead must belong to user's account
    lead = verify_tenant_access(db, Lead, lead_id, current_user)
    
    # Also verify lead_base and lote belong to same tenant
    base_nombre = None
    if lead.lead_base_id:
        base = db.query(LeadBase.nombre).filter(
            LeadBase.id == lead.lead_base_id,
            LeadBase.cuenta_id == current_user.cuenta_id
        ).first()
        if base:
            base_nombre = base.nombre

    lote_nombre = None
    if lead.lote_id:
        lote = db.query(Lote.nombre).filter(
            Lote.id == lead.lote_id,
            Lote.cuenta_id == current_user.cuenta_id
        ).first()
        if lote:
            lote_nombre = lote.nombre

    return _lead_to_dict(lead, base_nombre, lote_nombre)


# ---------------------------------------------------------------------------
# Bulk update template download
# ---------------------------------------------------------------------------
@router.get(
    "/accounts/{account_id}/leads/update-template",
    summary="Download Excel template for bulk lead update",
)
def download_update_template(
    account_id: uuid.UUID,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    from app.models.field import CustomField

    fields = (
        db.query(CustomField.nombre_campo)
        .filter(CustomField.cuenta_id == account_id)
        .order_by(CustomField.created_at)
        .all()
    )
    field_names = [f[0] for f in fields]

    wb = Workbook()
    ws = wb.active
    ws.title = "Actualizar Leads"
    ws.append(["id_lead"] + field_names)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"actualizar_leads_{account.nombre}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Bulk update via Excel
# ---------------------------------------------------------------------------
@router.post(
    "/accounts/{account_id}/leads/bulk-update",
    response_model=BulkUpdateResponse,
    summary="Bulk update leads from Excel file",
)
def bulk_update_leads(
    account_id: uuid.UUID,
    file: UploadFile,
    db: Session = Depends(get_db),
) -> dict:
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    try:
        content = file.file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel must have a header row and at least one data row")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    # First column MUST be id_lead
    if not header or header[0].lower() != "id_lead":
        raise HTTPException(
            status_code=400,
            detail="First column must be 'id_lead'",
        )

    update_columns = header[1:]
    if not update_columns:
        raise HTTPException(status_code=400, detail="No data columns to update (only id_lead found)")

    updated = 0
    not_found_ids: list[int] = []
    errors: list[str] = []
    leads_to_sync: list[tuple] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        raw_id = row[0] if row else None
        if raw_id is None or str(raw_id).strip() == "":
            continue

        try:
            lead_id_val = int(raw_id)
        except (ValueError, TypeError):
            errors.append(f"Row {row_idx}: invalid id_lead '{raw_id}'")
            continue

        lead = db.query(Lead).filter(
            Lead.cuenta_id == account_id,
            Lead.id_lead == lead_id_val,
        ).first()

        if not lead:
            not_found_ids.append(lead_id_val)
            continue

        # Build update dict from non-empty cells
        new_datos = dict(lead.datos)
        changed = False
        for col_idx, col_name in enumerate(update_columns):
            if not col_name:
                continue
            cell_val = row[col_idx + 1] if col_idx + 1 < len(row) else None
            if cell_val is not None and str(cell_val).strip() != "":
                new_datos[col_name] = cell_val
                changed = True

        if changed:
            lead.datos = new_datos
            # Also update the record
            if lead.record:
                lead.record.datos = new_datos
            updated += 1
            leads_to_sync.append((lead.id, new_datos))

    # Sync custom-field values into real columns
    try:
        field_rows = (
            db.query(CustomField.nombre_campo, CustomField.column_name)
            .filter(CustomField.cuenta_id == account_id, CustomField.column_name.isnot(None))
            .all()
        )
        field_map = {name: col for name, col in field_rows}
        if field_map:
            for lead_id, datos in leads_to_sync:
                sync_lead_columns(db, lead_id, datos, field_map)
    except Exception as e:
        logger.error("Failed to sync lead columns during bulk update: %s", e)

    db.commit()

    logger.info(
        "Bulk update for account %s: %d updated, %d not found",
        account_id, updated, len(not_found_ids),
    )

    return {
        "updated": updated,
        "not_found_ids": not_found_ids,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# GDPR / Data compliance
# ---------------------------------------------------------------------------

@router.get(
    "/leads/{lead_id}/export",
    summary="Export all data for a lead (GDPR right of access)",
)
def export_lead_data(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> JSONResponse:
    """Return all stored data for a lead as JSON (GDPR right of access)."""
    lead = verify_tenant_access(db, Lead, lead_id, current_user)

    export = {
        "lead_id": str(lead.id),
        "id_lead": lead.id_lead,
        "cuenta_id": str(lead.cuenta_id),
        "created_at": lead.created_at.isoformat() if lead.created_at else None,
        "updated_at": lead.updated_at.isoformat() if lead.updated_at else None,
        "datos": lead.datos,
        "score": lead.score,
        "temperatura": lead.temperatura,
        "assigned_to": str(lead.assigned_to) if lead.assigned_to else None,
    }

    filename = f"lead_{lead.id_lead}_export.json"
    return JSONResponse(
        content=export,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete(
    "/leads/{lead_id}/forget",
    summary="Anonymize lead PII (GDPR right to be forgotten)",
    status_code=200,
)
def forget_lead(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """
    Anonymize personally identifiable information for a lead.
    Preserves the lead record for analytics but removes PII.
    """
    lead = verify_tenant_access(db, Lead, lead_id, current_user)

    _PII_FIELDS = {"nombre", "apellido", "email", "correo", "telefono", "phone",
                   "tel", "celular", "dni", "rut", "direccion", "address"}

    anonymized = dict(lead.datos)
    fields_cleared = []
    for field in _PII_FIELDS:
        if field in anonymized:
            anonymized[field] = "[ANONIMIZADO]"
            fields_cleared.append(field)

    lead.datos = anonymized
    db.commit()

    logger.info(
        "Lead %s anonymized by user %s — fields: %s",
        lead_id, current_user.id, fields_cleared,
    )

    return {"success": True, "fields_anonymized": fields_cleared}
